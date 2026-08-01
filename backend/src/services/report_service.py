import os
from datetime import UTC, datetime

# openpyxl imports
import openpyxl

# ReportLab imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.alert import Alert
from models.node import Node

REPORTS_DIR = os.path.join(os.getcwd(), "reports_storage")
os.makedirs(REPORTS_DIR, exist_ok=True)


def utc_now() -> datetime:
    return datetime.now(UTC)


async def generate_pdf_report(db: AsyncSession, report_type: str = "weekly") -> tuple[str, str]:
    """
    Generates Executive PDF Report with Uptime Summary, Asset Inventory, and Incident History.
    Returns (filename, filepath).
    """
    now = utc_now()
    timestamp_str = now.strftime("%Y%m%d_%H%M%S")
    filename = f"report_{report_type}_{timestamp_str}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)

    # Fetch Data
    nodes_res = await db.execute(select(Node))
    nodes = list(nodes_res.scalars().all())

    alerts_res = await db.execute(select(Alert).order_by(Alert.triggered_at.desc()).limit(20))
    alerts = list(alerts_res.scalars().all())

    # Build PDF Document
    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    styles = getSampleStyleSheet()

    # Custom Title Style
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=12,
    )

    story.append(Paragraph(f"Executive Infrastructure Monitoring Report ({report_type.capitalize()})", title_style))
    story.append(Paragraph(f"Generated at: {now.strftime('%Y-%m-%d %H:%M:%S UTC')} | Platform: Infrastructure Monitoring & Auto-Topology", subtitle_style))
    story.append(Spacer(1, 10))

    # Executive Summary Table
    up_count = sum(1 for n in nodes if n.status == "up")
    down_count = sum(1 for n in nodes if n.status == "down")
    warn_count = sum(1 for n in nodes if n.status == "warning")
    total_nodes = len(nodes)
    sla_percentage = (up_count / total_nodes * 100) if total_nodes > 0 else 100.0

    summary_data = [
        ["Total Assets", "Nodes UP", "Nodes Warning", "Nodes DOWN", "Availability SLA"],
        [str(total_nodes), str(up_count), str(warn_count), str(down_count), f"{sla_percentage:.2f}%"],
    ]

    t_summary = Table(summary_data, colWidths=[100, 100, 100, 100, 140])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 16))

    # Section 1: Asset Inventory
    heading_style = ParagraphStyle("Heading2Custom", parent=styles["Heading2"], fontSize=14, spaceAfter=8, textColor=colors.HexColor("#1e293b"))
    story.append(Paragraph("Asset Inventory Summary", heading_style))

    inv_data = [["Name", "Type", "Status", "IP Address", "CPU Cores", "RAM (GB)"]]
    for n in nodes[:15]:
        inv_data.append([
            n.name,
            n.type.replace("_", " ").title(),
            n.status.upper(),
            n.ip_address or "N/A",
            str(n.cpu_cores or "N/A"),
            f"{round(n.ram_mb / 1024)}" if hasattr(n, "ram_mb") and n.ram_mb else "N/A",
        ])

    t_inv = Table(inv_data, colWidths=[120, 110, 70, 110, 65, 65])
    t_inv.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#334155")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(t_inv)
    story.append(Spacer(1, 16))

    # Section 2: Recent Incident History
    story.append(Paragraph("Recent Alert Incident Log", heading_style))
    alert_data = [["Severity", "Status", "Message", "Triggered At"]]
    for a in alerts[:10]:
        alert_data.append([
            a.severity.upper(),
            a.status.upper(),
            a.message[:45] + ("..." if len(a.message) > 45 else ""),
            a.triggered_at.strftime("%Y-%m-%d %H:%M"),
        ])

    if len(alerts) == 0:
        alert_data.append(["N/A", "NONE", "No recent incident alerts recorded.", "-"])

    t_alert = Table(alert_data, colWidths=[75, 75, 260, 130])
    t_alert.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#475569")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(t_alert)

    doc.build(story)
    return filename, filepath


async def generate_excel_report(db: AsyncSession, report_type: str = "weekly") -> tuple[str, str]:
    """
    Generates Excel Worksheets (.xlsx) Report with Inventory, Alerts, and SLA statistics.
    Returns (filename, filepath).
    """
    now = utc_now()
    timestamp_str = now.strftime("%Y%m%d_%H%M%S")
    filename = f"report_{report_type}_{timestamp_str}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = openpyxl.Workbook()

    # Sheet 1: Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"

    ws_summary.append(["Infrastructure Monitoring Executive Summary"])
    ws_summary.append(["Report Type:", report_type.capitalize()])
    ws_summary.append(["Generated At:", now.strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws_summary.append([])

    nodes_res = await db.execute(select(Node))
    nodes = list(nodes_res.scalars().all())

    up_count = sum(1 for n in nodes if n.status == "up")
    down_count = sum(1 for n in nodes if n.status == "down")
    warn_count = sum(1 for n in nodes if n.status == "warning")
    total_nodes = len(nodes)

    ws_summary.append(["Metric Name", "Value"])
    ws_summary.append(["Total Monitored Assets", total_nodes])
    ws_summary.append(["Nodes UP", up_count])
    ws_summary.append(["Nodes WARNING", warn_count])
    ws_summary.append(["Nodes DOWN", down_count])
    ws_summary.append(["Estimated Availability SLA", f"{((up_count/total_nodes*100) if total_nodes > 0 else 100):.2f}%"])

    # Sheet 2: Asset Inventory
    ws_inventory = wb.create_sheet(title="Asset Inventory")
    ws_inventory.append(["ID", "Name", "Type", "Status", "Review Status", "IP Address", "OS", "CPU Cores", "RAM (MB)", "Disk (GB)"])
    for n in nodes:
        ws_inventory.append([
            str(n.id),
            n.name,
            n.type,
            n.status,
            n.review_status,
            n.ip_address or "",
            n.os or "",
            n.cpu_cores or "",
            n.ram_mb or "",
            n.disk_gb or "",
        ])

    # Sheet 3: Alert History Log
    alerts_res = await db.execute(select(Alert).order_by(Alert.triggered_at.desc()).limit(100))
    alerts = list(alerts_res.scalars().all())

    ws_alerts = wb.create_sheet(title="Incident History Log")
    ws_alerts.append(["ID", "Node ID", "Severity", "Status", "Message", "Triggered At", "Resolved At", "Acknowledged By"])
    for a in alerts:
        ws_alerts.append([
            str(a.id),
            str(a.node_id),
            a.severity,
            a.status,
            a.message,
            a.triggered_at.strftime("%Y-%m-%d %H:%M:%S") if a.triggered_at else "",
            a.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if a.resolved_at else "",
            a.acknowledged_by or "",
        ])

    wb.save(filepath)
    return filename, filepath
